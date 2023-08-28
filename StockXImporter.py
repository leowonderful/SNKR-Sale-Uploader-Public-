import openpyxl
from datetime import datetime
import google.auth
from googleapiclient.discovery import build
from google.oauth2 import service_account
import os
import time

minCheck = 0

class itemSale: 
    def __init__(self, SKU, price, date, name, size, minPrice):
        self.SKU = SKU
        self.price = price
        self.date = date
        self.name = name
        self.size = size
        self.minPrice = minPrice



choose = int(input("Enter 0 for StockX imports, 1 for GOAT imports. "))


if choose == 0:
    minCheck = 1
    fname = input("What is the filename of the sheet for StockX? ")

    #Enter the path of your StockX spreadsheet.
    mySheet = openpyxl.load_workbook('Your-StockX-Spreadsheet-Path' + fname + '.xlsx')
    getSheet = mySheet['Sheet1']

    year = int(input("Which year to read from?"))
    month = int(input("What month to read from?"))
    day = int(input("What day to read from?"))

    storedEntries = []
    numOfEntriesAdded = 0

    for row in getSheet.iter_rows(min_row=2, values_only = True):
        if(row[0]):
            myRow = str(row[10])
            myRow = myRow[0:10]
            curRow = datetime.strptime(myRow, '%Y-%m-%d') 
            if(curRow >= datetime(year, month, day)):
                # print(curRow)
                newSale = itemSale(row[3],row[5],curRow,row[1],row[2], row[7])
                if row[3] == "N/A": #some clothing items have no SKU
                    newSale.SKU = row[1]
                storedEntries.append(newSale)
                numOfEntriesAdded+=1
                print("There are now" + str(numOfEntriesAdded) + "Entries.\n")

if choose == 1:
    fname = input("What is the filename of the sheet for GOAT?")

    #Enter the path of the GOAT Group spreadsheet.
    mySheet = openpyxl.load_workbook('Your-GOAT/Alias-Spreadsheet-Path' + fname + '.xlsx')
    getSheet = mySheet['completedsales']

    year = int(input("Which year to read from?"))
    month = int(input("What month to read from?"))
    day = int(input("What day to read from?"))

    storedEntries = []
    numOfEntriesAdded = 0

    for row in getSheet.iter_rows(min_row = 2, values_only = True): 
        if(row[0]):
            myRow = str(row[4])
            myRow = myRow[0:10]
            print(myRow)
            curRow = datetime.strptime(myRow, '%Y-%m-%d')
            if(curRow >= datetime(year,month,day)):
                newSale = itemSale(row[1],row[6],curRow,row[3],row[2],-1)
                storedEntries.append(newSale)
                numOfEntriesAdded+=1
                print("There are now" + str(numOfEntriesAdded) + "Entries.\n")


#Replace with your spreadsheet ID of choosing

SSID = 'your-spreadsheet-id' 

#Replace with your Google service credentials in JSON format. This must be the same folder as this python file or enter the full path.

creds = service_account.Credentials.from_service_account_file('Your-JSON-Google-Service-Credentials', scopes=['https://www.googleapis.com/auth/spreadsheets'])
service = build('sheets', 'v4', credentials=creds)

entriesInSheet = [99999,2,2,2,2,2,2,2,2,2,2,2]

#Enter your current StockX seller fee, excluding the 2.9% processing fee but including any rate discounts. 
sxPercent = [99999, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6, 6]

for entry in storedEntries:
    strMonth = entry.date.strftime("%B")
    numMonth = int(entry.date.strftime("%m"))
    DATERANGE = strMonth + '!A' + str(entriesInSheet[numMonth])
    SKURANGE = strMonth + '!B' + str(entriesInSheet[numMonth])
    SIZERANGE = strMonth + '!C' +  str(entriesInSheet[numMonth])
    PRICERANGE = strMonth + '!E' + str(entriesInSheet[numMonth])
    TYPERANGE = strMonth + '!F' + str(entriesInSheet[numMonth])

    myDate = entry.date.strftime("%m/%d/%y")
    mySKU = str(entry.SKU)
    myPrice = str(entry.price)
    mySize = str(entry.size)

    CASH_GIFT = "Cash/Gift"
    STOCKX_MAX = "StockX Max"
    GOAT = "GOAT"
    TYPE = "ERR"

    if(choose == 0):
        TYPE = STOCKX_MAX
    if(choose == 1):
        TYPE = GOAT
        myPrice = str(int(myPrice) / 100)
        mySKU = mySKU.replace(" ", "-")
    if(minCheck):
        if(sxPercent[numMonth] == 7):
            if(int(myPrice) < 129 or entry.minPrice / int(myPrice) > 0.901):
                #print("For debug the pct was " + str(entry.minPrice / int(myPrice)))
                myPrice = str(entry.minPrice)
                TYPE = CASH_GIFT
        if(sxPercent[numMonth] == 6):
            if(int(myPrice) < 150 or entry.minPrice / int(myPrice) > 0.911):
                #print("For debug the pct was " + str(entry.minPrice / int(myPrice)))
                myPrice = str(entry.minPrice)
                TYPE = CASH_GIFT

    values = [
            {
                'range': DATERANGE,
                'values': [
                    [myDate] #Date in MMDDYYYY format
                ],
            'majorDimension': 'ROWS'
            },
            {
             'range': SKURANGE,
             'values': [
                 [mySKU] #SKU which may need to be manually corrected. 
                ],
                'majorDimension': 'ROWS'
            },
            {
                'range': PRICERANGE,
                'values': [
                    [myPrice]
                ],
                'majorDimension': 'ROWS'
            },
            {
                'range': SIZERANGE,
                'values': [
                    [mySize]
                ],
                'majorDimension': 'ROWS'
            },
            {
                'range': TYPERANGE,
                'values': [
                    [TYPE]
                ],
                'majorDimension': 'ROWS'
            }
        ]

    body = {
        'data': values,
        'valueInputOption': 'USER_ENTERED'
    }

    result = service.spreadsheets().values().batchUpdate(
        spreadsheetId = SSID,
        body=body
    ).execute()
    print("successful update on " + str(entriesInSheet[numMonth]))
    entriesInSheet[numMonth] += 1 #Align the writing of entries to sheet correctly

    time.sleep(1) #avoid overloading the Sheets API; I will update this to properly support batch updates later
